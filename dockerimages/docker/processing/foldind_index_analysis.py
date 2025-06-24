import os
import math
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, NamedStyle, PatternFill, Alignment

# Diccionario de traducciones
traducciones = {
    "lh_caudalanteriorcingulate_foldind": "Corteza cingulada anterior caudal izquierda",
    "lh_caudalmiddlefrontal_foldind": "Corteza frontal media caudal izquierda",
    "lh_cuneus_foldind": "Cuneo izquierdo",
    "lh_entorhinal_foldind": "Entorrinal izquierdo",
    "lh_fusiform_foldind": "Fusiforme izquierdo",
    "lh_inferiorparietal_foldind": "Parietal inferior izquierdo",
    "lh_inferiortemporal_foldind": "Temporal inferior izquierdo",
    "lh_isthmuscingulate_foldind": "Istmo cingulado izquierdo",
    "lh_lateraloccipital_foldind": "Occipital lateral izquierdo",
    "lh_lateralorbitofrontal_foldind": "Orbitofrontal lateral izquierdo",
    "lh_lingual_foldind": "Lingual izquierdo",
    "lh_medialorbitofrontal_foldind": "Orbitofrontal medial izquierdo",
    "lh_middletemporal_foldind": "Temporal medio izquierdo",
    "lh_parahippocampal_foldind": "Parahipocampal izquierdo",
    "lh_paracentral_foldind": "Paracentral izquierdo",
    "lh_parsopercularis_foldind": "Pars opercularis izquierda",
    "lh_parsorbitalis_foldind": "Pars orbitalis izquierda",
    "lh_parstriangularis_foldind": "Pars triangularis izquierda",
    "lh_pericalcarine_foldind": "Pericalcarino izquierdo",
    "lh_postcentral_foldind": "Postcentral izquierdo",
    "lh_posteriorcingulate_foldind": "Corteza cingulada posterior izquierda",
    "lh_precentral_foldind": "Precentral izquierdo",
    "lh_precuneus_foldind": "Precuneo izquierdo",
    "lh_rostralanteriorcingulate_foldind": "Corteza cingulada anterior rostral izquierda",
    "lh_rostralmiddlefrontal_foldind": "Corteza frontal media rostral izquierda",
    "lh_superiorfrontal_foldind": "Frontal superior izquierdo",
    "lh_superiorparietal_foldind": "Parietal superior izquierdo",
    "lh_superiortemporal_foldind": "Temporal superior izquierdo",
    "lh_supramarginal_foldind": "Supramarginal izquierdo",
    "lh_frontalpole_foldind": "Polo frontal izquierdo",
    "lh_temporalpole_foldind": "Polo temporal izquierdo",
    "lh_transversetemporal_foldind": "Temporal transverso izquierdo",
    "lh_insula_foldind": "Ínsula izquierda",
    "BrainSegVolNotVent": "Volumen segmentado del cerebro sin ventrículos",
    "eTIV": "Volumen intracraneal estimado",
    "rh_caudalanteriorcingulate_foldind": "Corteza cingulada anterior caudal derecha",
    "rh_caudalmiddlefrontal_foldind": "Corteza frontal media caudal derecha",
    "rh_cuneus_foldind": "Cuneo derecho",
    "rh_entorhinal_foldind": "Entorrinal derecho",
    "rh_fusiform_foldind": "Fusiforme derecho",
    "rh_inferiorparietal_foldind": "Parietal inferior derecho",
    "rh_inferiortemporal_foldind": "Temporal inferior derecho",
    "rh_isthmuscingulate_foldind": "Istmo cingulado derecho",
    "rh_lateraloccipital_foldind": "Occipital lateral derecho",
    "rh_lateralorbitofrontal_foldind": "Orbitofrontal lateral derecho",
    "rh_lingual_foldind": "Lingual derecho",
    "rh_medialorbitofrontal_foldind": "Orbitofrontal medial derecho",
    "rh_middletemporal_foldind": "Temporal medio derecho",
    "rh_parahippocampal_foldind": "Parahipocampal derecho",
    "rh_paracentral_foldind": "Paracentral derecho",
    "rh_parsopercularis_foldind": "Pars opercularis derecha",
    "rh_parsorbitalis_foldind": "Pars orbitalis derecha",
    "rh_parstriangularis_foldind": "Pars triangularis derecha",
    "rh_pericalcarine_foldind": "Pericalcarino derecho",
    "rh_postcentral_foldind": "Postcentral derecho",
    "rh_posteriorcingulate_foldind": "Corteza cingulada posterior derecha",
    "rh_precentral_foldind": "Precentral derecho",
    "rh_precuneus_foldind": "Precuneo derecho",
    "rh_rostralanteriorcingulate_foldind": "Corteza cingulada anterior rostral derecha",
    "rh_rostralmiddlefrontal_foldind": "Corteza frontal media rostral derecha",
    "rh_superiorfrontal_foldind": "Frontal superior derecho",
    "rh_superiorparietal_foldind": "Parietal superior derecho",
    "rh_superiortemporal_foldind": "Temporal superior derecho",
    "rh_supramarginal_foldind": "Supramarginal derecho",
    "rh_frontalpole_foldind": "Polo frontal derecho",
    "rh_temporalpole_foldind": "Polo temporal derecho",
    "rh_transversetemporal_foldind": "Temporal transverso derecho",
    "rh_insula_foldind": "Ínsula derecha",
    "BrainSegVolNotVent": "Volumen segmentado del cerebro sin ventrículos",
    "eTIV": "Volumen intracraneal estimado"
}

def truncar_numero(num, decimales=2):
    """Redondea hacia abajo un número a una cantidad específica de decimales, manejando NaN e infinitos."""
    if pd.isna(num) or math.isinf(num):
        return "NaN"
    multiplicador = 10 ** decimales
    return f"{int(float(num) * multiplicador) / multiplicador:.{decimales}f}"


def seleccionar_base_control_foldind(edad, genero):
    """
    Selecciona la base de datos control de índices de plegamiento según edad y género.
    """
    base_dir = "database/morfo_cerebral/indice_plegamiento/"

    if edad <= 18:
        grupo = "18_29"
    elif edad <= 29:
        grupo = "18_29"
    elif edad <= 44:
        grupo = "30_44"
    elif edad <= 60:
        grupo = "45_60"
    else:
        grupo = "45_60"

    genero = "femenino" if genero.lower() == "f" else "masculino"
    
    file_path_estadisticos_control_lh = os.path.join(base_dir, f"grupo_{grupo}_{genero}_aparc_lh_stats_foldind_Z_Scores_Robustos.xlsx")
    file_path_estadisticos_control_rh = os.path.join(base_dir, f"grupo_{grupo}_{genero}_aparc_rh_stats_foldind_Z_Scores_Robustos.xlsx")

    # Verificar existencia de los archivos
    if not os.path.exists(file_path_estadisticos_control_lh) or not os.path.exists(file_path_estadisticos_control_rh):
        raise RuntimeError(f"No se encontraron los archivos de base de control en: {base_dir}")

    return file_path_estadisticos_control_lh, file_path_estadisticos_control_rh

def comparar_foldind(df_paciente, df_control):
    """
    Compara los índices de plegamiento del paciente con los valores del grupo control.
    Calcula el Z-score robusto y resalta valores fuera del umbral.
    """
    resultados = pd.DataFrame()
    umbral = 3.5

    for region in df_paciente.index:
        if region in df_control.index:
            valor_paciente_foldind = df_paciente.loc[region].values[0]
            mediana_control = df_control.loc[region, 'Mediana']
            mad_control = df_control.loc[region, 'MAD']
            ic_99_bajo = truncar_numero(df_control.loc[region, 'IC_99%_Bajo'], 2)
            ic_99_alto = truncar_numero(df_control.loc[region, 'IC_99%_Alto'], 2)

            # Manejo de valores NaN e infinitos
            if pd.isna(valor_paciente_foldind) or pd.isna(mediana_control) or pd.isna(mad_control) or mad_control == 0:
                z_score_paciente = "NaN"
            else:
                z_score_paciente = truncar_numero((0.6745 * (float(valor_paciente_foldind) - mediana_control) / mad_control), 2)

            # Determinar si está fuera del umbral
            fuera_umbral = '↑' if float(z_score_paciente) > umbral else '↓' if float(z_score_paciente) < -umbral else '✓'

            fila = pd.DataFrame({
                'Measure:foldind': [region],
                'Valor_Paciente': [truncar_numero(valor_paciente_foldind, 2)],
                'Z_Score_Paciente': [z_score_paciente],
                'IC_99%_Bajo': [ic_99_bajo],
                'IC_99%_Alto': [ic_99_alto],
                'Dentro_de_Umbral_±' + str(umbral): [fuera_umbral]
            })
            resultados = pd.concat([resultados, fila], ignore_index=True)

    return resultados

def procesar_foldind(stats_folder, edad, genero):
    """
    Procesa los índices de plegamiento cortical comparándolos con la base de datos de controles.
    """
    # Obtener archivos de índices de plegamiento del paciente
    file_path_paciente_lh = os.path.join(stats_folder, "aparc_lh_stats_foldind.txt")
    file_path_paciente_rh = os.path.join(stats_folder, "aparc_rh_stats_foldind.txt")

    # Seleccionar base de datos control
    file_path_estadisticos_control_lh, file_path_estadisticos_control_rh = seleccionar_base_control_foldind(edad, genero)

    # Leer archivos del paciente y del grupo control
    df_paciente_lh = pd.read_csv(file_path_paciente_lh, sep='\t', index_col='lh.aparc.foldind').drop(['lh_bankssts_foldind'], errors='ignore')
    df_paciente_rh = pd.read_csv(file_path_paciente_rh, sep='\t', index_col='rh.aparc.foldind').drop(['rh_bankssts_foldind'], errors='ignore')
    df_control_lh = pd.read_excel(file_path_estadisticos_control_lh, index_col='Measure:foldind', engine='openpyxl').drop(['lh_bankssts_foldind'], errors='ignore')
    df_control_rh = pd.read_excel(file_path_estadisticos_control_rh, index_col='Measure:foldind', engine='openpyxl').drop(['rh_bankssts_foldind'], errors='ignore')

    # Comparar índices de plegamiento del paciente con el grupo control
    resultados_lh = comparar_foldind(df_paciente_lh, df_control_lh)
    resultados_rh = comparar_foldind(df_paciente_rh, df_control_rh)
   
    # Añadir traducciones a los resultados
    resultados_lh['Regiones_ESP'] = resultados_lh['Measure:foldind'].map(traducciones)
    resultados_rh['Regiones_ESP'] = resultados_rh['Measure:foldind'].map(traducciones)

    # Guardar resultados en Excel
    output_path = os.path.join(stats_folder, "aparc_stats_foldind_Z_score_robusto.xlsx")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        resultados_lh.to_excel(writer, sheet_name='LH', index=False)
        resultados_rh.to_excel(writer, sheet_name='RH', index=False)
        workbook = writer.book

        # Estilo Arial tamaño 10
        arial_10 = NamedStyle(name="arial_10", font=Font(name="Arial", size=10))
        workbook.add_named_style(arial_10)

        # Formatear la salida
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

        for sheet_name in ['LH', 'RH']:
            worksheet = writer.sheets[sheet_name]
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.style = arial_10
                z_score_paciente = row[2].value
                ic_99_bajo = row[3].value
                ic_99_alto = row[4].value
                fuera_umbral = row[5].value

                if float(z_score_paciente) < float(ic_99_bajo) or float(z_score_paciente) > float(ic_99_alto):
                    for cell in row:
                        cell.fill = orange_fill

                if fuera_umbral in ['↑', '↓']:
                    for cell in row:
                        cell.fill = yellow_fill

    print(f"Resultados de índices de plegamiento guardados en {output_path}")
    print("")
