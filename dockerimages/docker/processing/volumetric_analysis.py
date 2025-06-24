#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import os
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, PatternFill


# Traducciones de regiones
traducciones = {
    "Left-Lateral-Ventricle": "Ventrículo lateral izquierdo",
    "Left-Inf-Lat-Vent": "Ventrículo inf. lat. izquierdo",
    "Left-Cerebellum-White-Matter": "Sustancia blanca del cerebelo izquierdo",
    "Left-Thalamus": "Tálamo izquierdo",
    "Left-Cerebellum-Cortex": "Corteza del cerebelo izquierdo",
    "Left-Caudate": "Núcleo caudado izquierdo",
    "Left-Putamen": "Putamen izquierdo",
    "Left-Pallidum": "Pálido izquierdo",
    "3rd-Ventricle": "3er ventrículo",
    "4th-Ventricle": "4to ventrículo",
    "Brain-Stem": "Tronco encefálico",
    "Left-Hippocampus": "Hipocampo izquierdo",
    "Left-Amygdala": "Amígdala izquierda",
    "CSF": "Líquido cefalorraquídeo",
    "Left-Accumbens-area": "Área del núcleo accumbens izquierdo",
    "Left-VentralDC": "DC ventral izquierdo",
    "Left-vessel": "Vaso sanguíneo izquierdo",
    "Left-choroid-plexus": "Plexo coroideo izquierdo",
    "Right-Lateral-Ventricle": "Ventrículo lateral derecho",
    "Right-Inf-Lat-Vent": "Ventrículo inf. lat. derecho",
    "Right-Cerebellum-White-Matter": "Sustancia blanca del cerebelo derecho",
    "Right-Cerebellum-Cortex": "Corteza del cerebelo derecho",
    "Right-Thalamus": "Tálamo derecho",
    "Right-Caudate": "Núcleo caudado derecho",
    "Right-Putamen": "Putamen derecho",
    "Right-Pallidum": "Pálido derecho",
    "Right-Hippocampus": "Hipocampo derecho",
    "Right-Amygdala": "Amígdala derecha",
    "Right-Accumbens-area": "Área del núcleo accumbens derecho",
    "Right-VentralDC": "DC ventral derecho",
    "Right-vessel": "Vaso sanguíneo derecho",
    "Right-choroid-plexus": "Plexo coroideo derecho",
    "5th-Ventricle": "5to ventrículo",
    "WM-hypointensities": "Hipointensidades de la sustancia blanca",
    "Left-WM-hypointensities": "Hipointensidades de la sustancia blanca izquierda",
    "Right-WM-hypointensities": "Hipointensidades de la sustancia blanca derecha",
    "non-WM-hypointensities": "Hipointensidades fuera de la sustancia blanca",
    "Left-non-WM-hypointensities": "Hipointensidades fuera de la sustancia blanca izquierda",
    "Right-non-WM-hypointensities": "Hipointensidades fuera de la sustancia blanca derecha",
    "Optic-Chiasm": "Quiasma óptico",
    "CC_Posterior": "Cuerpo calloso posterior",
    "CC_Mid_Posterior": "Cuerpo calloso medio posterior",
    "CC_Central": "Cuerpo calloso central",
    "CC_Mid_Anterior": "Cuerpo calloso medio anterior",
    "CC_Anterior": "Cuerpo calloso anterior",
    "BrainSegVol": "Volumen segmentado del cerebro",
    "BrainSegVolNotVent": "Volumen segmentado del cerebro sin ventrículos",
    "lhCortexVol": "Volumen de la corteza izquierda",
    "rhCortexVol": "Volumen de la corteza derecha",
    "CortexVol": "Volumen total de la corteza",
    "lhCerebralWhiteMatterVol": "Sustancia blanca cerebral izquierda",
    "rhCerebralWhiteMatterVol": "Sustancia blanca cerebral derecha",
    "CerebralWhiteMatterVol": "Sustancia blanca total",
    "SubCortGrayVol": "Sustancia gris subcortical",
    "TotalGrayVol": "Sustancia gris total",
    "SupraTentorialVol": "Volumen supratentorial",
    "SupraTentorialVolNotVent": "Volumen supratentorial sin ventrículos",
    "MaskVol": "Volumen de máscara",
    "BrainSegVol-to-eTIV": "Volumen del segmento cerebral respecto a eTIV",
    "MaskVol-to-eTIV": "Volumen de máscara respecto a eTIV",
    "lhSurfaceHoles": "Huecos en la superficie izquierda",
    "rhSurfaceHoles": "Huecos en la superficie derecha",
    "SurfaceHoles": "Huecos en la superficie",
    "EstimatedTotalIntraCranialVol": "Volumen Intracraneal Total estimado"
}

# Diccionario de pares de regiones homólogas
pares_regiones = {
    "Left-Lateral-Ventricle": "Right-Lateral-Ventricle",
    "Left-Inf-Lat-Vent": "Right-Inf-Lat-Vent",
    "Left-Cerebellum-White-Matter": "Right-Cerebellum-White-Matter",
    "Left-Cerebellum-Cortex": "Right-Cerebellum-Cortex",
    "Left-Thalamus": "Right-Thalamus",
    "Left-Caudate": "Right-Caudate",
    "Left-Putamen": "Right-Putamen",
    "Left-Pallidum": "Right-Pallidum",
    "Left-Hippocampus": "Right-Hippocampus",
    "Left-Amygdala": "Right-Amygdala",
    "Left-Accumbens-area": "Right-Accumbens-area",
    "Left-choroid-plexus": "Right-choroid-plexus",
    "lhCortexVol": "rhCortexVol",
    "lhCerebralWhiteMatterVol": "rhCerebralWhiteMatterVol"
}

# Diccionario de traducción de regiones
traduccion_regiones = {
    "Lateral-Ventricle": "Ventrículo Lateral",
    "Inf-Lat-Vent": "Ventrículo Inferior Lateral",
    "Cerebellum-White-Matter": "Sustancia Blanca del Cerebelo",
    "Cerebellum-Cortex": "Corteza del Cerebelo",
    "Thalamus": "Tálamo",
    "Caudate": "Caudado",
    "Putamen": "Putamen",
    "Pallidum": "Pálido",
    "Hippocampus": "Hipocampo",
    "Amygdala": "Amígdala",
    "Accumbens-area": "Área Accumbens",
    "choroid-plexus": "Plexo coroideo",
    "CortexVol": "Corteza Cerebral",
    "CerebralWhiteMatterVol": "Sustancia Blanca Cerebral"
}
    
def seleccionar_base_control(edad, genero):
    """
    Selecciona automáticamente la base de datos control según la edad y el género.
    """
    # Ruta absoluta al directorio que contiene las bases de datos
    base_dir = "database/morfo_cerebral/volumen/"

    if edad <= 18:
        grupo = "18_29"
    elif edad <= 29:
        grupo = "18_29"
    elif edad <= 44:
        grupo = "30_44"
    elif edad <= 60:
        grupo = "45_60"
    else:
        grupo = "45_60"

    genero = "femenino" if genero.lower() == "f" else "masculino"
    archivo_base = f"grupo_{grupo}_{genero}_aseg_stats_etiv_IC_Bootstrap.xlsx"

    # Combinar base_dir con el archivo
    archivo_path = os.path.join(base_dir, archivo_base)
    
    # Verificar que el archivo existe
    if not os.path.exists(archivo_path):
        raise RuntimeError(f"No se encontró el archivo de base de control en: {archivo_path}")

    return archivo_path
    
    


def truncar_numero(num, decimales=2):
    multiplicador = 10 ** decimales
    return math.trunc(num * multiplicador) / multiplicador

def normalizar_nombre_region(nombre):
    if nombre == "Left-Thalamus-Proper":
        return "Left-Thalamus"
    elif nombre == "Right-Thalamus-Proper":
        return "Right-Thalamus"
    return nombre


def calcular_asimetria(left_vol, right_vol):
    diferencia_absoluta = abs(left_vol - right_vol)
    promedio_volumenes = (left_vol + right_vol) / 2
    return (diferencia_absoluta / promedio_volumenes) * 100 if promedio_volumenes != 0 else 0
    

def calcular_asimetrias_pares(df_final, base_control_path):
    """
    Itera sobre los pares de regiones y calcula las asimetrías.
    Extrae los IC_99% de la hoja 'Asimetrias' del archivo base de datos.
    """
    # Leer la hoja 'Asimetrias' del archivo base de datos
    df_asimetrias_estadisticas = pd.read_excel(base_control_path, sheet_name='Asimetrias', engine='openpyxl')

    # Crear DataFrame con columnas correctas
    resultados_asimetria = pd.DataFrame(columns=['Region', 'Asimetria', 'IC_99%_Bajo', 'IC_99%_Alto', 'Rango_normal_ajustado_por_edad_según_AIP'])

    # Iterar sobre cada par de regiones homólogas
    for left_region, right_region in pares_regiones.items():
        left_vol = df_final.loc[df_final['Measure:volume'] == left_region, 'Volumen_cm3'].values
        right_vol = df_final.loc[df_final['Measure:volume'] == right_region, 'Volumen_cm3'].values

        if left_vol.size > 0 and right_vol.size > 0:
            # Calcular asimetría
            asimetria = calcular_asimetria(float(left_vol[0]), float(right_vol[0]))
            region_base = left_region.replace('Left-', '').replace('lh', '')
            region_traducida = traduccion_regiones.get(region_base, region_base)

            # Buscar los IC_99% en la hoja 'Asimetrias'
            rango_control = df_asimetrias_estadisticas.loc[
                df_asimetrias_estadisticas['Region'] == region_traducida, ['IC_99%_Bajo', 'IC_99%_Alto']
            ]

            if not rango_control.empty:
                ic_99_bajo, ic_99_alto = rango_control['IC_99%_Bajo'].values[0], rango_control['IC_99%_Alto'].values[0]
                rango_normal = f"{ic_99_bajo} - {ic_99_alto}"
            else:
                ic_99_bajo, ic_99_alto, rango_normal = "N/A", "N/A", "N/A"

            # Añadir resultados al DataFrame
            resultados_asimetria = pd.concat([resultados_asimetria, pd.DataFrame({
                'Region': [region_traducida],
                'Asimetria': [truncar_numero(asimetria, 2)],
                'IC_99%_Bajo': [ic_99_bajo],
                'IC_99%_Alto': [ic_99_alto],
                'Rango_normal_ajustado_por_edad_según_AIP': [rango_normal]
            })], ignore_index=True)

    return resultados_asimetria
    print("")

def procesar_volumenes(stats_folder, base_control_path):
    """
    Procesa los volúmenes y calcula las asimetrías a partir de los datos de FreeSurfer.
    Incluye 'Rango_normal_ajustado_por_edad_según_%VIT' en la hoja de volúmenes.
    Formatea 'Volumen_cm3' y 'Volumen_%VIT'.
    """
    # Leer archivos
    file_path_volumenes_cm3 = os.path.join(stats_folder, 'aseg_stats_cm3.txt')
    file_path_volumenes_porcentaje = os.path.join(stats_folder, 'aseg_stats_etiv.txt')

    df_volumenes_cm3 = pd.read_csv(file_path_volumenes_cm3, sep='\t', header=None,
                                   names=['Measure:volume', 'Volumen_cm3'], skiprows=1)
    df_volumenes_porcentaje = pd.read_csv(file_path_volumenes_porcentaje, sep='\t', header=None,
                                          names=['Measure:volume', 'Volumen_%VIT'], skiprows=1)
    df_control = pd.read_excel(base_control_path, sheet_name='Bootstrap_Results', engine='openpyxl')

    # Truncar IC_99% y IC_95% valores a dos decimales
    for column in ['IC_99%_Bajo', 'IC_99%_Alto', 'IC_95%_Bajo', 'IC_95%_Alto']:
        df_control[column] = df_control[column].apply(lambda x: truncar_numero(x, 2))

    # Normalizar nombres
    for df in [df_volumenes_cm3, df_volumenes_porcentaje, df_control]:
        df['Measure:volume'] = df['Measure:volume'].apply(normalizar_nombre_region)

    # Fusionar tablas
    df_merged = pd.merge(df_volumenes_cm3, df_volumenes_porcentaje, on='Measure:volume', how='inner')
    df_final = pd.merge(df_merged, 
                        df_control[['Measure:volume', 'IC_99%_Bajo', 'IC_99%_Alto', 'IC_95%_Bajo', 'IC_95%_Alto']], 
                        on='Measure:volume', how='inner')
    
    

    # Formatear 'Volumen_cm3' y 'Volumen_%VIT'
    df_final['Volumen_cm3'] = df_final['Volumen_cm3'].apply(lambda x: f"{truncar_numero(x, 2):.2f}")
    df_final['Volumen_%VIT'] = df_final['Volumen_%VIT'].apply(lambda x: f"{truncar_numero(x, 2):.2f}")

    # Agregar columna 'Rango_normal_ajustado_por_edad_según_%VIT'
    df_final['Rango_normal_ajustado_por_edad_según_%VIT'] = df_final['IC_99%_Bajo'].astype(str) + ' - ' + df_final['IC_99%_Alto'].astype(str)

    df_final['Regiones_ESP'] = df_final['Measure:volume'].map(traducciones)
    
    # Ordenar las columnas en el DataFrame final
    df_final = df_final[['Measure:volume', 'Volumen_cm3', 'Volumen_%VIT',
                     'Rango_normal_ajustado_por_edad_según_%VIT',
                     'IC_99%_Bajo', 'IC_99%_Alto', 'IC_95%_Bajo', 'IC_95%_Alto',
                     'Regiones_ESP']]

    # Calcular asimetrías usando la ruta de base_control_path
    resultados_asimetria = calcular_asimetrias_pares(df_final, base_control_path)

    return df_final, resultados_asimetria


    
def exportar_volumetria_excel(df_final, resultados_asimetria, output_path):
    """
    Exporta los datos de volúmenes y asimetrías a un archivo Excel.
    Resalta valores fuera del rango normal en la hoja de volúmenes y asimetrías.
    """
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Escribir hojas en el Excel
        df_final.to_excel(writer, sheet_name='Volumenes', index=False)
        resultados_asimetria.to_excel(writer, sheet_name='Asimetrias', index=False)

        workbook = writer.book
        estilo_normal = NamedStyle(name="arial_10", font=Font(name="Arial", size=10))

        # Hoja de Volumenes
        ws_volumenes = writer.sheets['Volumenes']
        for row in ws_volumenes.iter_rows(min_row=2, max_row=ws_volumenes.max_row, min_col=1, max_col=ws_volumenes.max_column):
            for cell in row:
                cell.style = estilo_normal

        # Incluir IC_95 en el archivo final
        if 'IC_95%_Bajo' in df_final.columns and 'IC_95%_Alto' in df_final.columns:
            for idx, row in df_final.iterrows():
                try:
                    vol = float(row['Volumen_%VIT'])
                    ic_95_bajo = float(row['IC_95%_Bajo'])
                    ic_95_alto = float(row['IC_95%_Alto'])
                    ic_99_bajo = float(row['IC_99%_Bajo'])
                    ic_99_alto = float(row['IC_99%_Alto'])

                    # Resaltar filas fuera del rango IC_95 en negrita
                    if vol < ic_95_bajo or vol > ic_95_alto:
                        for col in range(len(row)):
                            cell = ws_volumenes.cell(row=idx + 2, column=col + 1)
                            cell.font = Font(name="Arial", size=10, bold=True)

                    # Resaltar filas fuera del rango IC_99 en color anaranjado
                    if vol < ic_99_bajo or vol > ic_99_alto:
                        for col in range(len(row)):
                            cell = ws_volumenes.cell(row=idx + 2, column=col + 1)
                            cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

                except ValueError:
                    continue

        # Hoja de Asimetrias
        ws_asimetria = writer.sheets['Asimetrias']
        if 'IC_99%_Bajo' in resultados_asimetria.columns and 'IC_99%_Alto' in resultados_asimetria.columns:
            for idx, row in resultados_asimetria.iterrows():
                try:
                    asimetria = float(row['Asimetria'])
                    ic_bajo = float(row['IC_99%_Bajo']) if row['IC_99%_Bajo'] != "N/A" else None
                    ic_alto = float(row['IC_99%_Alto']) if row['IC_99%_Alto'] != "N/A" else None

                    # Resaltar filas fuera del rango IC_99 en color anaranjado
                    if ic_bajo is not None and (asimetria < ic_bajo or asimetria > ic_alto):
                        for col in range(1, len(row) + 1):
                            cell = ws_asimetria.cell(row=idx + 2, column=col)
                            cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

                except ValueError:
                    continue

    print(f"Archivo Excel exportado en {output_path}")
    



