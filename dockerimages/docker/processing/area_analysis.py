# %%
import os
import math
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, NamedStyle, PatternFill, Alignment

# Diccionario de traducciones
traducciones = {
    "lh_bankssts_area": "Bancos del surco temporal superior izquierdo",
    "lh_caudalanteriorcingulate_area": "Corteza cingulada anterior caudal izquierda",
    "lh_caudalmiddlefrontal_area": "Corteza frontal media caudal izquierda",
    "lh_cuneus_area": "Cuneo izquierdo",
    "lh_entorhinal_area": "Entorrinal izquierda",
    "lh_fusiform_area": "Fusiforme izquierda",
    "lh_inferiorparietal_area": "Parietal inferior izquierda",
    "lh_inferiortemporal_area": "Temporal inferior izquierda",
    "lh_isthmuscingulate_area": "Istmo cingulado izquierdo",
    "lh_lateraloccipital_area": "Occipital lateral izquierda",
    "lh_lateralorbitofrontal_area": "Orbitofrontal lateral izquierda",
    "lh_lingual_area": "Lingual izquierda",
    "lh_medialorbitofrontal_area": "Orbitofrontal medial izquierda",
    "lh_middletemporal_area": "Temporal media izquierda",
    "lh_parahippocampal_area": "Parahipocampal izquierda",
    "lh_paracentral_area": "Paracentral izquierda",
    "lh_parsopercularis_area": "Pars opercularis izquierda",
    "lh_parsorbitalis_area": "Pars orbitalis izquierda",
    "lh_parstriangularis_area": "Pars triangularis izquierda",
    "lh_pericalcarine_area": "Pericalcarina izquierda",
    "lh_postcentral_area": "Postcentral izquierda",
    "lh_posteriorcingulate_area": "Corteza cingulada posterior izquierda",
    "lh_precentral_area": "Precentral izquierda",
    "lh_precuneus_area": "Precuneo izquierdo",
    "lh_rostralanteriorcingulate_area": "Corteza cingulada anterior rostral izquierda",
    "lh_rostralmiddlefrontal_area": "Corteza frontal media rostral izquierda",
    "lh_superiorfrontal_area": "Frontal superior izquierda",
    "lh_superiorparietal_area": "Parietal superior izquierda",
    "lh_superiortemporal_area": "Temporal superior izquierda",
    "lh_supramarginal_area": "Supramarginal izquierda",
    "lh_frontalpole_area": "Polo frontal izquierdo",
    "lh_temporalpole_area": "Polo temporal izquierdo",
    "lh_transversetemporal_area": "Temporal transversa izquierda",
    "lh_insula_area": "Ínsula izquierda",
    "lh_WhiteSurfArea_area": "Superficie de sustancia blanca izquierda",
    "BrainSegVolNotVent": "Volumen segmentado del cerebro sin ventrículos",
    "eTIV": "Volumen intracraneal estimado",
    "rh_bankssts_area": "Bancos del surco temporal superior derecho",
    "rh_caudalanteriorcingulate_area": "Corteza cingulada anterior caudal derecha",
    "rh_caudalmiddlefrontal_area": "Corteza frontal media caudal derecha",
    "rh_cuneus_area": "Cuneo derecho",
    "rh_entorhinal_area": "Entorrinal derecha",
    "rh_fusiform_area": "Fusiforme derecha",
    "rh_inferiorparietal_area": "Parietal inferior derecha",
    "rh_inferiortemporal_area": "Temporal inferior derecha",
    "rh_isthmuscingulate_area": "Istmo cingulado derecho",
    "rh_lateraloccipital_area": "Occipital lateral derecha",
    "rh_lateralorbitofrontal_area": "Orbitofrontal lateral derecha",
    "rh_lingual_area": "Lingual derecha",
    "rh_medialorbitofrontal_area": "Orbitofrontal medial derecha",
    "rh_middletemporal_area": "Temporal media derecha",
    "rh_parahippocampal_area": "Parahipocampal derecha",
    "rh_paracentral_area": "Paracentral derecha",
    "rh_parsopercularis_area": "Pars opercularis derecha",
    "rh_parsorbitalis_area": "Pars orbitalis derecha",
    "rh_parstriangularis_area": "Pars triangularis derecha",
    "rh_pericalcarine_area": "Pericalcarina derecha",
    "rh_postcentral_area": "Postcentral derecha",
    "rh_posteriorcingulate_area": "Corteza cingulada posterior derecha",
    "rh_precentral_area": "Precentral derecha",
    "rh_precuneus_area": "Precuneo derecho",
    "rh_rostralanteriorcingulate_area": "Corteza cingulada anterior rostral derecha",
    "rh_rostralmiddlefrontal_area": "Corteza frontal media rostral derecha",
    "rh_superiorfrontal_area": "Frontal superior derecha",
    "rh_superiorparietal_area": "Parietal superior derecha",
    "rh_superiortemporal_area": "Temporal superior derecha",
    "rh_supramarginal_area": "Supramarginal derecha",
    "rh_frontalpole_area": "Polo frontal derecho",
    "rh_temporalpole_area": "Polo temporal derecho",
    "rh_transversetemporal_area": "Temporal transversa derecha",
    "rh_insula_area": "Ínsula derecha",
    "rh_WhiteSurfArea_area": "Superficie de sustancia blanca derecha",
    "BrainSegVolNotVent": "Volumen segmentado del cerebro sin ventrículos",
    "eTIV": "Volumen intracraneal estimado"
}

def truncar_numero(num, decimales=2):
    """Trunca un número a una cantidad específica de decimales."""
    multiplicador = 10 ** decimales
    return f"{math.trunc(num * multiplicador) / multiplicador:.{decimales}f}"

def seleccionar_base_control_area(edad, genero):
    """
    Selecciona la base de datos control de areas según edad y género.
    """
    base_dir = "database/morfo_cerebral/area_superficie_cortical"
    
    if edad <= 18:
        grupo = "18_29"
    elif edad <= 29:
        grupo = "18_29"
    elif edad <= 44:
        grupo = "30_44"
    elif edad <= 60:
        grupo = "45_60"
    else:
        grupo = "45_60"

    genero = "femenino" if genero.lower() == "f" else "masculino"
    
    file_path_estadisticos_control_lh = os.path.join(base_dir, f"grupo_{grupo}_{genero}_aparc_lh_stats_area_Z_Scores_Robustos.xlsx")
    file_path_estadisticos_control_rh = os.path.join(base_dir, f"grupo_{grupo}_{genero}_aparc_rh_stats_area_Z_Scores_Robustos.xlsx")

    # Verificar existencia de los archivos
    if not os.path.exists(file_path_estadisticos_control_lh) or not os.path.exists(file_path_estadisticos_control_rh):
        raise RuntimeError(f"No se encontraron los archivos de base de control en: {base_dir}")

    return file_path_estadisticos_control_lh, file_path_estadisticos_control_rh

def comparar_areas(df_paciente, df_control):
    """
    Compara las areas corticales del paciente con los valores del grupo control.
    Calcula el Z-score y resalta valores fuera del umbral.
    """
    resultados = pd.DataFrame()
    umbral = 3.5

    for region in df_paciente.index:
        if region in df_control.index:
            valor_paciente_mm2 = df_paciente.loc[region].values[0]
            mediana_control = df_control.loc[region, 'Mediana']
            mad_control = df_control.loc[region, 'MAD']
            ic_99_bajo = truncar_numero(df_control.loc[region, 'IC_99%_Bajo'], 2)
            ic_99_alto = truncar_numero(df_control.loc[region, 'IC_99%_Alto'], 2)

            # Convertir el valor del paciente a Z-score robusto y truncar
            z_score_paciente = truncar_numero((0.6745 * (float(valor_paciente_mm2) - mediana_control) / mad_control), 2)
            
            # Excluir BrainSegVolNotVent y eTIV de ciertas columnas
            if region in ["BrainSegVolNotVent", "eTIV"]:
                fila = pd.DataFrame({
                    'Measure:area': [region],
                    'Valor_Paciente_mm2': [valor_paciente_mm2],
                    'Z_Score_Paciente': [None],
                    'IC_99%_Bajo': [None],
                    'IC_99%_Alto': [None],
                    'Dentro_de_Umbral_±' + str(umbral): [None]
                })
            else:
                fila = pd.DataFrame({
                    'Measure:area': [region],
                    'Valor_Paciente_mm2': [truncar_numero(valor_paciente_mm2, 2)],
                    'Z_Score_Paciente': [z_score_paciente],
                    'IC_99%_Bajo': [ic_99_bajo],
                    'IC_99%_Alto': [ic_99_alto],
                    'Dentro_de_Umbral_±' + str(umbral): ['↑' if float(z_score_paciente) > umbral else '↓' if float(z_score_paciente) < -umbral else '✓']
                })
            resultados = pd.concat([resultados, fila], ignore_index=True)
    return resultados

def procesar_areas(stats_folder, edad, genero):
    """
    Procesa las areas corticales comparándolos con la base de datos de controles.
    """
    # Obtener archivos de areas del paciente
    file_path_paciente_lh = os.path.join(stats_folder, "aparc_lh_stats_area.txt")
    file_path_paciente_rh = os.path.join(stats_folder, "aparc_rh_stats_area.txt")

    # Seleccionar base de datos control
    file_path_estadisticos_control_lh, file_path_estadisticos_control_rh = seleccionar_base_control_area(edad, genero)

    # Leer archivos del paciente y del grupo control
    df_paciente_lh = pd.read_csv(file_path_paciente_lh, sep='\t', index_col='lh.aparc.area')
    df_paciente_rh = pd.read_csv(file_path_paciente_rh, sep='\t', index_col='rh.aparc.area')
    df_control_lh = pd.read_excel(file_path_estadisticos_control_lh, index_col='Measure:area', engine='openpyxl')
    df_control_rh = pd.read_excel(file_path_estadisticos_control_rh, index_col='Measure:area', engine='openpyxl')

    # Comparar areas del paciente con el grupo control
    resultados_lh = comparar_areas(df_paciente_lh, df_control_lh)
    resultados_rh = comparar_areas(df_paciente_rh, df_control_rh)
    
    # Añadir traducciones a los resultados
    resultados_lh['Regiones_ESP'] = resultados_lh['Measure:area'].map(traducciones)
    resultados_rh['Regiones_ESP'] = resultados_rh['Measure:area'].map(traducciones)

    # Guardar resultados en Excel
    output_path = os.path.join(stats_folder, "aparc_stats_area_Z_score_robusto.xlsx")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        resultados_lh.to_excel(writer, sheet_name='LH', index=False)
        resultados_rh.to_excel(writer, sheet_name='RH', index=False)
        workbook = writer.book

        # Estilo Arial tamaño 10
        arial_10 = NamedStyle(name="arial_10", font=Font(name="Arial", size=10))
        workbook.add_named_style(arial_10)

        # Formatear la salida
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        center_alignment = Alignment(horizontal='center')

        for sheet_name in ['LH', 'RH']:
            worksheet = writer.sheets[sheet_name]

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.style = arial_10

                region_name = row[0].value
                z_score_paciente = row[2].value
                ic_99_bajo = row[3].value
                ic_99_alto = row[4].value
                fuera_umbral = row[5].value  # Valor en la columna "Fuera_Umbral"

                # Excluir las regiones "BrainSegVolNotVent" y "eTIV" de las comparaciones
                if region_name not in ["BrainSegVolNotVent", "eTIV"]:
                    # Condición para aplicar color naranja si está fuera del IC99%
                    if float(z_score_paciente) < float(ic_99_bajo) or float(z_score_paciente) > float(ic_99_alto):
                        for cell in row:
                            cell.fill = orange_fill

                    # Condición para aplicar color amarillo si está fuera del umbral
                    if fuera_umbral in ['↑', '↓']:
                        for cell in row:
                            cell.fill = yellow_fill

                # Alinear al centro los valores de las columnas relevantes
                row[1].alignment = center_alignment
                row[2].alignment = center_alignment
                row[3].alignment = center_alignment
                row[4].alignment = center_alignment
                row[5].alignment = center_alignment
    
    print(f"Resultados de areas guardados en {output_path}")
    print("")




