import os
import math
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, NamedStyle, PatternFill, Alignment

# Diccionario de traducciones
traducciones = {
    "lh_bankssts_thickness": "Bancos del surco temporal superior izquierdo",
    "lh_caudalanteriorcingulate_thickness": "Corteza cingulada anterior caudal izquierda",
    "lh_caudalmiddlefrontal_thickness": "Corteza frontal media caudal izquierda",
    "lh_cuneus_thickness": "Cuneo izquierdo",
    "lh_entorhinal_thickness": "Entorrinal izquierdo",
    "lh_fusiform_thickness": "Fusiforme izquierdo",
    "lh_inferiorparietal_thickness": "Parietal inferior izquierdo",
    "lh_inferiortemporal_thickness": "Temporal inferior izquierdo",
    "lh_isthmuscingulate_thickness": "Istmo cingulado izquierdo",
    "lh_lateraloccipital_thickness": "Occipital lateral izquierdo",
    "lh_lateralorbitofrontal_thickness": "Orbitofrontal lateral izquierdo",
    "lh_lingual_thickness": "Lingual izquierdo",
    "lh_medialorbitofrontal_thickness": "Orbitofrontal medial izquierdo",
    "lh_middletemporal_thickness": "Temporal medio izquierdo",
    "lh_parahippocampal_thickness": "Parahipocampal izquierdo",
    "lh_paracentral_thickness": "Paracentral izquierdo",
    "lh_parsopercularis_thickness": "Pars opercularis izquierda",
    "lh_parsorbitalis_thickness": "Pars orbitalis izquierda",
    "lh_parstriangularis_thickness": "Pars triangularis izquierda",
    "lh_pericalcarine_thickness": "Pericalcarino izquierdo",
    "lh_postcentral_thickness": "Postcentral izquierdo",
    "lh_posteriorcingulate_thickness": "Corteza cingulada posterior izquierda",
    "lh_precentral_thickness": "Precentral izquierdo",
    "lh_precuneus_thickness": "Precuneo izquierdo",
    "lh_rostralanteriorcingulate_thickness": "Corteza cingulada anterior rostral izquierda",
    "lh_rostralmiddlefrontal_thickness": "Corteza frontal media rostral izquierda",
    "lh_superiorfrontal_thickness": "Frontal superior izquierdo",
    "lh_superiorparietal_thickness": "Parietal superior izquierdo",
    "lh_superiortemporal_thickness": "Temporal superior izquierdo",
    "lh_supramarginal_thickness": "Supramarginal izquierdo",
    "lh_frontalpole_thickness": "Polo frontal izquierdo",
    "lh_temporalpole_thickness": "Polo temporal izquierdo",
    "lh_transversetemporal_thickness": "Temporal transverso izquierdo",
    "lh_insula_thickness": "Ínsula izquierda",
    "lh_MeanThickness_thickness": "Espesor medio del hemisferio izquierdo",
    "BrainSegVolNotVent": "Volumen segmentado del cerebro sin ventrículos",
    "eTIV": "Volumen intracraneal estimado",
    "rh_bankssts_thickness": "Bancos del surco temporal superior derecho",
    "rh_caudalanteriorcingulate_thickness": "Corteza cingulada anterior caudal derecha",
    "rh_caudalmiddlefrontal_thickness": "Corteza frontal media caudal derecha",
    "rh_cuneus_thickness": "Cuneo derecho",
    "rh_entorhinal_thickness": "Entorrinal derecho",
    "rh_fusiform_thickness": "Fusiforme derecho",
    "rh_inferiorparietal_thickness": "Parietal inferior derecho",
    "rh_inferiortemporal_thickness": "Temporal inferior derecho",
    "rh_isthmuscingulate_thickness": "Istmo cingulado derecho",
    "rh_lateraloccipital_thickness": "Occipital lateral derecho",
    "rh_lateralorbitofrontal_thickness": "Orbitofrontal lateral derecho",
    "rh_lingual_thickness": "Lingual derecho",
    "rh_medialorbitofrontal_thickness": "Orbitofrontal medial derecho",
    "rh_middletemporal_thickness": "Temporal medio derecho",
    "rh_parahippocampal_thickness": "Parahipocampal derecho",
    "rh_paracentral_thickness": "Paracentral derecho",
    "rh_parsopercularis_thickness": "Pars opercularis derecha",
    "rh_parsorbitalis_thickness": "Pars orbitalis derecha",
    "rh_parstriangularis_thickness": "Pars triangularis derecha",
    "rh_pericalcarine_thickness": "Pericalcarino derecho",
    "rh_postcentral_thickness": "Postcentral derecho",
    "rh_posteriorcingulate_thickness": "Corteza cingulada posterior derecha",
    "rh_precentral_thickness": "Precentral derecho",
    "rh_precuneus_thickness": "Precuneo derecho",
    "rh_rostralanteriorcingulate_thickness": "Corteza cingulada anterior rostral derecha",
    "rh_rostralmiddlefrontal_thickness": "Corteza frontal media rostral derecha",
    "rh_superiorfrontal_thickness": "Frontal superior derecho",
    "rh_superiorparietal_thickness": "Parietal superior derecho",
    "rh_superiortemporal_thickness": "Temporal superior derecho",
    "rh_supramarginal_thickness": "Supramarginal derecho",
    "rh_frontalpole_thickness": "Polo frontal derecho",
    "rh_temporalpole_thickness": "Polo temporal derecho",
    "rh_transversetemporal_thickness": "Temporal transverso derecho",
    "rh_insula_thickness": "Ínsula derecha",
    "rh_MeanThickness_thickness": "Espesor medio del hemisferio derecho",
    "BrainSegVolNotVent": "Volumen segmentado del cerebro sin ventrículos",
    "eTIV": "Volumen intracraneal estimado"
}

def truncar_numero(num, decimales=2):
    """Trunca un número a una cantidad específica de decimales."""
    multiplicador = 10 ** decimales
    return f"{math.trunc(num * multiplicador) / multiplicador:.{decimales}f}"

def seleccionar_base_control_espesores(edad, genero):
    """
    Selecciona la base de datos control de espesores según edad y género.
    """
    base_dir = "database/morfo_cerebral/espesor_cortical/"

    if edad <= 18:
        grupo = "18_29"
    elif edad <= 29:
        grupo = "18_29"
    elif edad <= 44:
        grupo = "30_44"
    elif edad <= 60:
        grupo = "45_60"
    else:
        grupo = "45_60"

    genero = "femenino" if genero.lower() == "f" else "masculino"
    
    file_path_estadisticos_control_lh = os.path.join(base_dir, f"grupo_{grupo}_{genero}_aparc_lh_stats_thickness_Z_Scores_Robustos.xlsx")
    file_path_estadisticos_control_rh = os.path.join(base_dir, f"grupo_{grupo}_{genero}_aparc_rh_stats_thickness_Z_Scores_Robustos.xlsx")

    # Verificar existencia de los archivos
    if not os.path.exists(file_path_estadisticos_control_lh) or not os.path.exists(file_path_estadisticos_control_rh):
        raise RuntimeError(f"No se encontraron los archivos de base de control en: {base_dir}")

    return file_path_estadisticos_control_lh, file_path_estadisticos_control_rh

def comparar_espesores(df_paciente, df_control):
    """
    Compara los espesores corticales del paciente con los valores del grupo control.
    Calcula el Z-score robusto y resalta valores fuera del umbral.
    """
    resultados = pd.DataFrame()
    umbral = 3.5

    for region in df_paciente.index:
        if region in df_control.index:
            valor_paciente_mm = df_paciente.loc[region].values[0]
            mediana_control = df_control.loc[region, 'Mediana']
            mad_control = df_control.loc[region, 'MAD']
            ic_99_bajo = truncar_numero(df_control.loc[region, 'IC_99%_Bajo'], 2)
            ic_99_alto = truncar_numero(df_control.loc[region, 'IC_99%_Alto'], 2)

            # Calcular Z-score robusto
            z_score_paciente = truncar_numero((0.6745 * (float(valor_paciente_mm) - mediana_control) / mad_control), 2)

            # Determinar si está fuera del umbral
            fuera_umbral = '↑' if float(z_score_paciente) > umbral else '↓' if float(z_score_paciente) < -umbral else '✓'

            fila = pd.DataFrame({
                'Measure:thickness': [region],
                'Valor_Paciente_mm': [truncar_numero(valor_paciente_mm, 2)],
                'Z_Score_Paciente': [z_score_paciente],
                'IC_99%_Bajo': [ic_99_bajo],
                'IC_99%_Alto': [ic_99_alto],
                'Dentro_de_Umbral_±' + str(umbral): [fuera_umbral]
            })
            resultados = pd.concat([resultados, fila], ignore_index=True)

    return resultados

def procesar_espesores(stats_folder, edad, genero):
    """
    Procesa los espesores corticales comparándolos con la base de datos de controles.
    """
    # Obtener archivos de espesores del paciente
    file_path_paciente_lh = os.path.join(stats_folder, "aparc_lh_stats_thickness.txt")
    file_path_paciente_rh = os.path.join(stats_folder, "aparc_rh_stats_thickness.txt")

    # Seleccionar base de datos control
    file_path_estadisticos_control_lh, file_path_estadisticos_control_rh = seleccionar_base_control_espesores(edad, genero)

    # Leer archivos del paciente y del grupo control
    df_paciente_lh = pd.read_csv(file_path_paciente_lh, sep='\t', index_col='lh.aparc.thickness')
    df_paciente_rh = pd.read_csv(file_path_paciente_rh, sep='\t', index_col='rh.aparc.thickness')
    df_control_lh = pd.read_excel(file_path_estadisticos_control_lh, index_col='Measure:thickness', engine='openpyxl')
    df_control_rh = pd.read_excel(file_path_estadisticos_control_rh, index_col='Measure:thickness', engine='openpyxl')

    # Comparar espesores del paciente con el grupo control
    resultados_lh = comparar_espesores(df_paciente_lh, df_control_lh)
    resultados_rh = comparar_espesores(df_paciente_rh, df_control_rh)
    
    # Añadir traducciones a los resultados
    resultados_lh['Regiones_ESP'] = resultados_lh['Measure:thickness'].map(traducciones)
    resultados_rh['Regiones_ESP'] = resultados_rh['Measure:thickness'].map(traducciones)

    # Guardar resultados en Excel
    output_path = os.path.join(stats_folder, "aparc_stats_thickness_Z_score_robusto.xlsx")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        resultados_lh.to_excel(writer, sheet_name='LH', index=False)
        resultados_rh.to_excel(writer, sheet_name='RH', index=False)
        workbook = writer.book

        # Estilo Arial tamaño 10
        arial_10 = NamedStyle(name="arial_10", font=Font(name="Arial", size=10))
        workbook.add_named_style(arial_10)

        # Formatear la salida
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

        for sheet_name in ['LH', 'RH']:
            worksheet = writer.sheets[sheet_name]
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.style = arial_10
                z_score_paciente = row[2].value
                ic_99_bajo = row[3].value
                ic_99_alto = row[4].value
                fuera_umbral = row[5].value

                if float(z_score_paciente) < float(ic_99_bajo) or float(z_score_paciente) > float(ic_99_alto):
                    for cell in row:
                        cell.fill = orange_fill

                if fuera_umbral in ['↑', '↓']:
                    for cell in row:
                        cell.fill = yellow_fill
    
    print(f"Resultados de espesores guardados en {output_path}")
    print("")
